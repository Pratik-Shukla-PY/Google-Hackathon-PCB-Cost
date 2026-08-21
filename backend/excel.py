import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

def generate_bom_excel(
    project: Dict[str, Any],
    bom_rows: List[Dict[str, Any]],
    detections: List[Dict[str, Any]],
    output_path: str
):
    wb = Workbook()
    
    # Define styles
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_normal = Font(name="Segoe UI", size=10)
    font_link = Font(name="Segoe UI", size=10, color="0000FF", underline="single")
    font_title = Font(name="Segoe UI", size=12, bold=True)
    
    # Header Fill (Slate 100)
    fill_header = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    # Generic Row Fill (Light Amber/Yellow for warnings)
    fill_generic = PatternFill(start_color="FFFDF0", end_color="FFFDF0", fill_type="solid")
    
    # Border
    thin_side = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # ----------------------------------------------------
    # SHEET 1: Cost BOM
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Cost BOM"
    ws1.views.sheetView[0].showGridLines = True
    
    headers1 = [
        "Line No", "Class", "Designators", "Qty", "Marking Text", "Package",
        "Measured Length (mm)", "Measured Width (mm)", "Package Status",
        "Manufacturer Name", "Manufacturer Part Number", "Description",
        "Distributor Name", "Distributor Part Number", "Datasheet URL",
        "Product Page URL", "Match Basis", "Unit Price", "Extended Cost",
        "Price Break Qty", "MOQ", "Stock Status", "Price Date",
        "Confidence Level", "Mfr Read", "Part No Read", "Price Read", "Note"
    ]
    
    # Write Headers
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all
    
    # Write Rows
    current_row = 2
    for r in bom_rows:
        row_values = [
            r["line_number"],
            r["component_class"],
            r["designators"],
            r["quantity"],
            r["marking_text"],
            r["package"],
            r["measured_length"],
            r["measured_width"],
            r["package_resolution_status"],
            r["manufacturer"],
            r["part_number"],
            r["description"],
            r["distributor"],
            r["distributor_part_number"],
            r["datasheet_url"],  # will convert to hyperlink below
            r["product_page_url"],  # will convert to hyperlink below
            r["match_basis"],
            r["unit_price"],
            r["extended_cost"],
            r["price_break_qty"],
            r["moq"],
            r["stock_status"],
            r["price_date"],
            r["confidence"],
            r["manufacturer_read"],
            r["part_number_read"],
            r["price_read"],
            r["note"]
        ]
        
        is_generic = r["match_basis"].lower() == "generic"
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws1.cell(row=current_row, column=col_idx)
            cell.font = font_normal
            cell.border = border_all
            
            # Shading generic rows
            if is_generic:
                cell.fill = fill_generic
                
            # Alignment & formatting
            if col_idx in [1, 4, 7, 8, 20, 21]:  # numeric integer columns
                cell.value = val
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [18, 19]:  # currency columns
                cell.value = val
                cell.number_format = '"$"#,##0.000'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [15, 16] and val:  # hyperlink URLs
                cell.value = "Link"
                cell.hyperlink = val
                cell.font = font_link
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = val
                cell.alignment = Alignment(horizontal="left")
                
        current_row += 1
        
    # Sum/Total Row
    total_row_idx = current_row
    ws1.cell(row=total_row_idx, column=1, value="Total").font = font_bold
    ws1.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="left")
    ws1.cell(row=total_row_idx, column=1).border = border_all
    
    # Empty cells border
    for col in range(2, 29):
        ws1.cell(row=total_row_idx, column=col).border = border_all
        
    # Extended Cost Sum formula
    cost_col_letter = get_column_letter(19)
    sum_cell = ws1.cell(row=total_row_idx, column=19, value=f"=SUM({cost_col_letter}2:{cost_col_letter}{total_row_idx-1})")
    sum_cell.font = font_bold
    sum_cell.number_format = '"$"#,##0.00'
    sum_cell.alignment = Alignment(horizontal="right")
    
    ws1.freeze_panes = "A2"
    
    # ----------------------------------------------------
    # SHEET 2: Detections (Audit Trail)
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Detections")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["Designator", "Class", "Package", "Measured Width (mm)", "Measured Height (mm)", "Confidence", "Global Bounding Box"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_all
        
    current_row = 2
    for det in detections:
        bbox_str = f"xmin: {int(det['bbox']['xmin'])}, ymin: {int(det['bbox']['ymin'])}, xmax: {int(det['bbox']['xmax'])}, ymax: {int(det['bbox']['ymax'])}"
        row_values = [
            det.get("designator") or "Anon",
            det["component_class"],
            det.get("package") or "Unknown",
            det.get("measured_width"),
            det.get("measured_height"),
            det["confidence"],
            bbox_str
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws2.cell(row=current_row, column=col_idx)
            cell.font = font_normal
            cell.border = border_all
            cell.value = val
            if col_idx in [4, 5]:
                cell.alignment = Alignment(horizontal="right")
                if val:
                    cell.number_format = '0.00'
            else:
                cell.alignment = Alignment(horizontal="left")
        current_row += 1
        
    ws2.freeze_panes = "A2"
    
    # ----------------------------------------------------
    # SHEET 3: Run Summary (Metadata)
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Run Summary")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=1, column=1, value="PCB Teardown Costing Summary").font = font_title
    
    # Calculate sum of extended cost
    total_cost = sum(r["extended_cost"] or 0 for r in bom_rows)
    generic_count = sum(1 for r in bom_rows if r["match_basis"].lower() == "generic")
    total_count = sum(r["quantity"] for r in bom_rows)
    
    summary_data = [
        ("Image Filename", project["filename"]),
        ("Upload Timestamp", project["created_at"]),
        ("Scale Factor (px/mm)", f"{project['scale_factor']:.3f}" if project["scale_factor"] else "Uncalibrated"),
        ("Reference Box (pixels)", f"x: {int(project['reference_box']['box_x'])}, y: {int(project['reference_box']['box_y'])}, w: {int(project['reference_box']['box_w'])}, h: {int(project['reference_box']['box_h'])}" if project["reference_box"] else "N/A"),
        ("Reference Box Size (mm)", f"{project['reference_box']['real_w']} x {project['reference_box']['real_h']} mm" if project["reference_box"] else "N/A"),
        ("Build Volume Selected", project["build_volume"]),
        ("Tile Grid Used", "3x3 tiles, 15% overlap"),
        ("Model Version", "Claude 3.5 Sonnet (Vision) / Gemini 2.5 Flash"),
        ("Total Unique Line Items", len(bom_rows)),
        ("Total Component Count", total_count),
        ("Total Board Cost (at volume)", f"${total_cost:.3f}"),
        ("Generically Costed Items", generic_count),
    ]
    
    # Write summary metadata
    row_idx = 3
    for label, val in summary_data:
        cell_lbl = ws3.cell(row=row_idx, column=1, value=label)
        cell_lbl.font = font_bold
        cell_lbl.border = border_all
        cell_lbl.fill = fill_header
        
        cell_val = ws3.cell(row=row_idx, column=2, value=val)
        cell_val.font = font_normal
        cell_val.border = border_all
        row_idx += 1
        
    # Exclusion statement
    row_idx += 2
    ws3.cell(row=row_idx, column=1, value="BOM Exclusions & Liability Disclaimer:").font = font_bold
    row_idx += 1
    
    disclaimer = (
        "This teardown cost model covers only the components visible on this side of the board. "
        "It excludes: (1) the bare printed circuit board fabrication cost, (2) assembly labor, "
        "manufacturing, and placement fees, (3) testing, inspection, and yield losses, "
        "(4) mechanical packaging, enclosure housings, and heat sinks, "
        "(5) any components covered under metallic shielding cans or mounted on the reverse side of the board. "
        "All values marked as 'generic' are costed using industry commodity pricing benchmarks for standard package sizes and are not guaranteed to match specific designer selections."
    )
    
    cell_disc = ws3.cell(row=row_idx, column=1, value=disclaimer)
    cell_disc.font = font_normal
    cell_disc.alignment = Alignment(wrap_text=True, vertical="top")
    ws3.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx+5, end_column=3)
    
    # Auto-adjust column widths across all sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            # Handle merged cell check to prevent calculation errors
            vals = [cell.value for cell in col if cell.value and not type(cell).__name__ == 'MergedCell']
            max_len = max(len(str(v or '')) for v in vals) if vals else 5
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    wb.save(output_path)
